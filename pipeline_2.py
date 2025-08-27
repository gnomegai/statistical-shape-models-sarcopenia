#!/usr/bin/env python
# coding: utf-8

# ### 1. Initialisation

# In[1]:


import os
import shutil
import glob
import subprocess
import time
import itertools
import numpy as np
import pandas as pd
from sklearn.decomposition import PCA
from tqdm import tqdm

# Création du fichier Excel
import matplotlib
matplotlib.use("Agg") 
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment
import openpyxl

# Couleurs du Terminal
from colorama import Fore, Style, init as colorama_init
colorama_init(autoreset=True)

# ShapeWorks 
import shapeworks as sw


# ##### Paramètres Généraux

# In[ ]:


DATASET_NAME       = "D_J_128_TA"
DATASET_PATHS      = [
        # KEEP THIS COMMENTS
        #('./DATA/RF_FULGUR_M', 'RF'),
        #('./DATA/RF_FULGUR_PRED', 'RFP'),
        #('./DATA/RF_DIASEM', 'RFDIA'),
        #('./DATA/RF_FULGUR_SAMPLE', 'TEST1'),
        #('./DATA/RF_FULGUR_SAMPLE_2', 'TEST2'),
        #('./DATA/RF_DIASEM_SAMPLE', 'TESTDIA'),
        #('C:/Users/gossa/OneDrive/Documents/Travail INSA Lyon/PFE/BDD/FULGUR_label4_RF_post_process/FULGUR_label4_RF_post_process', 'PRGM')
    # ('/home/jupyter-gossard/Code_PFE/BDD/healthy_subj_JEREMIE_RF/RF_JEREMIE_Aligned', 'Jeunes'),
    # ('/home/jupyter-gossard/Code_PFE/BDD/label_RF_DIASEM/RF_Diasem_Aligned', 'Malades')
    #('/home/jupyter-gossard/Code_PFE/BDD/FULGUR_label4_RF_post_process', 'Sportif'),
    # ('/home/jupyter-gossard/Code_PFE/BDD/VL_diasem_labels/labels/manual_gz/resampled', 'Malades'),
    # ('/home/jupyter-gossard/Code_PFE/BDD/VL_young_labels/labels/manual_gz/resampled', 'Jeunes')
    # ('/home/jupyter-gossard/Code_PFE/BDD/VL_young_labels/labels/manual_gz/unique', 'Jeunes')
    
    ('/home/jupyter-gossard/Code_PFE/BDD/TA_diasem_labels/labels_recale', 'Malades'),
    ('/home/jupyter-gossard/Code_PFE/BDD/TA_young_labels/labels_recale', 'Jeunes')
    # ('/home/jupyter-gossard/Code_PFE/BDD/VL_young_labels/labels/unique', 'Jeunes')
]

# --- Paramètres Variables --- #

GRID_OPTIMIZATION = {
    "number_of_particles": [128], #"number_of_particles": [16, 32, 64, 128, 256, 512],
}

GRID_GROOMING = {
}


# ##### Paramètres par défaut

# In[14]:


# --- Paramètres de Grooming par défaut ---
ANTIALIAS_ITERATIONS = 30
ISO_SPACING          = [1, 1, 1]
PAD_SIZE             = 5
PAD_VALUE            = 0
ISO_VALUE            = 0.5
ICP_ITERATIONS       = 200

# --- Paramètres d'Optimisation par défaut ---
OPT_PARAMS = {
    "number_of_particles":       128,
    "use_normals":               1, ###### 0
    "normals_strength":          5.0,  ##### 10
    "checkpointing_interval":    1000,
    "keep_checkpoints":          0,
    "iterations_per_split":      1000,
    "optimization_iterations":   4000, ##### 1000
    "starting_regularization":   100, ### 100
    "ending_regularization":     1, #### 1
    "relative_weighting":        10, ### 1
    "initial_relative_weighting":0.1, ### 0.1
    "procrustes_interval":       0,
    "procrustes_scaling":        1,
    "save_init_splits":          0,
    "verbosity":                 0,
    "multiscale":                1,
    "multiscale_particles":      32,
    "tiny_test":                 False,
    "use_single_scale":          0,
    "reflect":                   False,
    "mesh_mode":                 True
}


# --- Dossiers ---

SHAPE_EXT          = ".nii.gz"
DT_EXT             = ".nrrd"
BASE_OUTPUT_DIR    = os.path.abspath(os.path.join(".", "OUTPUT_PIPELINE"))
OUTPUT_DB = os.path.abspath(os.path.join(".", "OUTPUT_DB"))

# Création du dossier ou se sauvegardent automatiquement les excel
os.makedirs(OUTPUT_DB, exist_ok=True)

# Si le dossier de sortie existe déjà:
# 1- On déplace les fichiers .xlsx dans le dossier OUTPUT_DB
# 2- On supprime tout le contenu du dossier de sortie
if os.path.exists(BASE_OUTPUT_DIR):
    for root, dirs, files in os.walk(BASE_OUTPUT_DIR):
        for file in files:
            if file.endswith(".xlsx"):
                source_file = os.path.join(root, file)
                new_filename = f"{os.path.splitext(file)[0]}_{int(time.time())}.xlsx"
                dest_file = os.path.join(OUTPUT_DB, new_filename)
                shutil.move(source_file, dest_file)
    shutil.rmtree(BASE_OUTPUT_DIR)

# Création du dossier de sortie
os.makedirs(BASE_OUTPUT_DIR, exist_ok=True)


# In[15]:


def color_print(msg, color=Fore.CYAN, style=Style.NORMAL):
    """ Print a colored message """
    print(color + style + msg + Style.RESET_ALL)


# ##### Aquisition des données

# In[16]:


def acquire_data(dataset_paths, shape_ext, output_path):
    color_print("\n--- Step 1. Acquire Data ---", Fore.GREEN, Style.BRIGHT)
    os.makedirs(output_path, exist_ok=True)

    shape_filenames = []
    dataset_ids = []

    for data_path, dataset_id in dataset_paths:
        files = sorted(glob.glob(os.path.join(data_path, '*' + shape_ext)))
        shape_filenames.extend(files)
        dataset_ids.extend([dataset_id] * len(files))

    color_print(f"  Nombre de shapes : {len(shape_filenames)}", Fore.YELLOW)
    return shape_filenames, dataset_ids


# In[17]:


def get_particles(particles_dir, particle_type="world"):
    particles = []
    names = []
    for filename in os.listdir(particles_dir):
        if filename.endswith(particle_type + ".particles"):
            data = np.loadtxt(os.path.join(particles_dir, filename))
            particles.append(data)
            names.append(os.path.splitext(filename)[0])
    if not particles:
        return None, None
    return np.array(particles), names


# ### 2. Preprocessing (Acquisition, Grooming, Rigid)

# In[18]:


def groom_shapes(shape_filenames, dataset_ids, groom_dir):
    color_print("\n--- Step 2. Groom - Data Pre-processing ---", Fore.GREEN, Style.BRIGHT)
    os.makedirs(groom_dir, exist_ok=True)
    
    start_groom = time.time()
    
    shape_seg_list = []
    shape_names = []

    for i, shape_filename in enumerate(tqdm(shape_filenames, desc="Grooming shapes")):
        # print(shape_filename)
        dataset_id = dataset_ids[i]
        base_shape_name = os.path.basename(shape_filename).replace(SHAPE_EXT, '')
        shape_name = f"{dataset_id}_{base_shape_name}"
        shape_names.append(shape_name)

        shape_seg = sw.Image(shape_filename)
        shape_seg_list.append(shape_seg)
        # print(shape_seg.dims())

        # bounding_box = sw.ImageUtils.boundingBox([shape_seg], ISO_VALUE).pad(2)
        # shape_seg.crop(bounding_box)
        #########
        # padding = 2
        # max_padding = 20  # valeur max de padding pour éviter une boucle infinie
        
        # while True:
        #     try:
        #         bounding_box = sw.ImageUtils.boundingBox([shape_seg], ISO_VALUE).pad(padding)
        #         print(f"Image size: {shape_seg.size()}")
        #         print(f"Image origin: {shape_seg.origin()}")
        #         print(f"Image spacing: {shape_seg.spacing()}")
        #         print(f"Bounding box min: {bounding_box.min}")
        #         print(f"Bounding box max: {bounding_box.max}")

        #         shape_seg.crop(bounding_box)
        #         break  # succès, on sort de la boucle
        #     except Exception as e:
        #         print(f"⚠️ Erreur avec padding={padding} : {e}")
        #         padding += 2
        #         if padding > max_padding:
        #             print("❌ Échec : padding trop grand, abandon.")
        #             break
        #########
        shape_seg.antialias(ANTIALIAS_ITERATIONS).resample(ISO_SPACING, sw.InterpolationType.Linear).binarize()
        shape_seg.pad(PAD_SIZE, PAD_VALUE)

    groom_time = time.time() - start_groom
    color_print(f"  Grooming time: {groom_time:.2f}s", Fore.YELLOW)

    return shape_seg_list, shape_names, groom_time




def rigid_transformations(shape_seg_list, shape_names, groom_dir):
    color_print("\n--- Step 3. Groom - Rigid Transformations ---", Fore.GREEN, Style.BRIGHT)
    os.makedirs(groom_dir, exist_ok=True)

    start_rigid = time.time()

    ref_index = sw.find_reference_image_index(shape_seg_list)
    ref_seg = shape_seg_list[ref_index]
    ref_name = shape_names[ref_index]

    ref_filename = os.path.join(groom_dir, 'reference' + DT_EXT)
    ref_seg.write(ref_filename)
    color_print(f"  Image de référence trouvée : {ref_name}", Fore.YELLOW)

    transform_dir = os.path.join(groom_dir, 'rigid_transforms')
    os.makedirs(transform_dir, exist_ok=True)

    rigid_transforms = []

    for shape_seg, shape_name in tqdm(zip(shape_seg_list, shape_names),
                                      desc="Calcul des transformations",
                                      total=len(shape_seg_list)):
        rigid_transform = shape_seg.createRigidRegistrationTransform(ref_seg, ISO_VALUE, ICP_ITERATIONS)
        rigid_transform = sw.utils.getVTKtransform(rigid_transform)
        rigid_transforms.append(rigid_transform)

        transform_filename = os.path.join(transform_dir, f'{shape_name}_to_{ref_name}_transform.txt')
        np.savetxt(transform_filename, rigid_transform)

        shape_seg.antialias(ANTIALIAS_ITERATIONS).computeDT(0).gaussianBlur(1.5)

    output_subdir = 'distance_transforms'
    output_dir = os.path.join(groom_dir, output_subdir)
    os.makedirs(output_dir, exist_ok=True)

    groomed_files = []
    for shape_seg, shape_name in zip(shape_seg_list, shape_names):
        out_name = os.path.join(output_dir, shape_name + DT_EXT)
        shape_seg.write(out_name)
        groomed_files.append(out_name)

    rigid_time = time.time() - start_rigid
    color_print(f"  Rigid time: {rigid_time:.2f}s", Fore.YELLOW)

    return rigid_transforms, groomed_files, rigid_time

def run_preprocessing(dataset_paths, shape_ext, grooming_params, base_dir):
    """
    Applique les grooming_params (PAD_SIZE, ICP_ITERATIONS, etc.) 
    => acquisition => grooming => rigid
    Retourne tout + le temps grooming + le temps rigid
    """
    global PAD_SIZE, ICP_ITERATIONS, ISO_SPACING, ANTIALIAS_ITERATIONS, PAD_VALUE, ISO_VALUE

    # Mise à jour des variables globales
    for k, v in grooming_params.items():
        if k == "PAD_SIZE":
            PAD_SIZE = v
        elif k == "ICP_ITERATIONS":
            ICP_ITERATIONS = v
        elif k == "ISO_SPACING":
            ISO_SPACING = v
        elif k == "ANTIALIAS_ITERATIONS":
            ANTIALIAS_ITERATIONS = v
        elif k == "PAD_VALUE":
            PAD_VALUE = v
        elif k == "ISO_VALUE":
            ISO_VALUE = v
        else:
            color_print(f"[WARNING] Paramètre grooming inconnu: {k} = {v}", Fore.RED)

    output_path = os.path.join(base_dir, "OUTPUT")
    os.makedirs(output_path, exist_ok=True)
    groom_dir = os.path.join(output_path, "groomed")

    shape_filenames, dataset_ids = acquire_data(dataset_paths, shape_ext, output_path)

    # Groom
    shape_seg_list, shape_names, t_groom = groom_shapes(shape_filenames, dataset_ids, groom_dir)

    # Rigid
    rigid_transforms, groomed_files, t_rigid = rigid_transformations(shape_seg_list, shape_names, groom_dir)

    return (shape_seg_list, shape_filenames, dataset_ids, shape_names, 
            rigid_transforms, groomed_files, t_groom, t_rigid)


# ### 3. Optimisation et Particules

# In[19]:


def optimize_particles(shape_seg_list, shape_filenames, rigid_transforms, groomed_files, output_path):
    color_print("\n--- Step 4. Optimize - Particle Based Optimization ---", Fore.GREEN, Style.BRIGHT)
    os.makedirs(output_path, exist_ok=True)

    domain_type, groomed_files_out = sw.data.get_optimize_input(
        groomed_files,
        OPT_PARAMS["mesh_mode"]
    )

    subjects = []
    for i in range(len(shape_seg_list)):
        subj = sw.Subject()
        subj.set_number_of_domains(1)

        subj.set_original_filenames([os.path.abspath(shape_filenames[i])])
        subj.set_groomed_filenames([os.path.abspath(groomed_files_out[i])])
        subj.set_groomed_transforms([rigid_transforms[i].flatten()])

        try:
            subj.set_domain_type(0, domain_type)
        except AttributeError:
            pass

        subjects.append(subj)

    project = sw.Project()
    project.set_subjects(subjects)
    parameters = sw.Parameters()

    valid_params = {
        "number_of_particles":        OPT_PARAMS["number_of_particles"],
        "use_normals":                OPT_PARAMS["use_normals"],
        "normals_strength":           OPT_PARAMS["normals_strength"],
        "checkpointing_interval":     OPT_PARAMS["checkpointing_interval"],
        "keep_checkpoints":           OPT_PARAMS["keep_checkpoints"],
        "iterations_per_split":       OPT_PARAMS["iterations_per_split"],
        "optimization_iterations":    OPT_PARAMS["optimization_iterations"],
        "starting_regularization":    OPT_PARAMS["starting_regularization"],
        "ending_regularization":      OPT_PARAMS["ending_regularization"],
        "relative_weighting":         OPT_PARAMS["relative_weighting"],
        "initial_relative_weighting": OPT_PARAMS["initial_relative_weighting"],
        "procrustes_interval":        OPT_PARAMS["procrustes_interval"],
        "procrustes_scaling":         OPT_PARAMS["procrustes_scaling"],
        "save_init_splits":           OPT_PARAMS["save_init_splits"],
        "verbosity":                  OPT_PARAMS["verbosity"]
    }

    if OPT_PARAMS.get("tiny_test", False):
        valid_params["number_of_particles"] = 32
        valid_params["optimization_iterations"] = 25

    if not OPT_PARAMS.get("use_single_scale", 0):
        valid_params["multiscale"] = 1
        valid_params["multiscale_particles"] = OPT_PARAMS["multiscale_particles"]

    for k, v in valid_params.items():
        parameters.set(k, sw.Variant([v]))

    project.set_parameters("optimize", parameters)

    proj_file = os.path.join(output_path, f"{DATASET_NAME}.swproj")
    project.save(proj_file)

    color_print("  Lancement de l'optimisation via shapeworks...", Fore.YELLOW)
    cmd = ['shapeworks', 'optimize', '--progress', '--name', proj_file]
    subprocess.check_call(cmd, cwd=output_path)

    args_for_check = type('ArgsForCheck', (object,), {})()
    args_for_check.tiny_test = OPT_PARAMS.get("tiny_test", False)
    args_for_check.verify    = False
    sw.utils.check_results(args_for_check, proj_file)

    return os.path.join(output_path, f"{DATASET_NAME}_particles")

    # # Nouvelle Fonction Optimisation avec l'ajout fichier mesh aligné

# def optimize_particles(shape_seg_list, shape_filenames, rigid_transforms, groomed_files, output_path):
#     color_print("\n--- Step 4. Optimize - Particle Based Optimization ---", Fore.GREEN, Style.BRIGHT)
#     os.makedirs(output_path, exist_ok=True)

#     domain_type, groomed_files_out = sw.data.get_optimize_input(
#         groomed_files,
#         OPT_PARAMS["mesh_mode"]
#     )

#     subjects = []
#     for i in range(len(shape_seg_list)):
#         subj = sw.Subject()
#         subj.set_number_of_domains(1)

#         subj.set_original_filenames([os.path.abspath(shape_filenames[i])])
#         subj.set_groomed_filenames([os.path.abspath(groomed_files_out[i])])
#         subj.set_groomed_transforms([rigid_transforms[i].flatten()])

#         try:
#             subj.set_domain_type(0, domain_type)
#         except AttributeError:
#             pass

#         subjects.append(subj)

#     project = sw.Project()
#     project.set_subjects(subjects)
#     parameters = sw.Parameters()

#     valid_params = {
#         "number_of_particles":        OPT_PARAMS["number_of_particles"],
#         "use_normals":                OPT_PARAMS["use_normals"],
#         "normals_strength":           OPT_PARAMS["normals_strength"],
#         "checkpointing_interval":     OPT_PARAMS["checkpointing_interval"],
#         "keep_checkpoints":           OPT_PARAMS["keep_checkpoints"],
#         "iterations_per_split":       OPT_PARAMS["iterations_per_split"],
#         "optimization_iterations":    OPT_PARAMS["optimization_iterations"],
#         "starting_regularization":    OPT_PARAMS["starting_regularization"],
#         "ending_regularization":      OPT_PARAMS["ending_regularization"],
#         "relative_weighting":         OPT_PARAMS["relative_weighting"],
#         "initial_relative_weighting": OPT_PARAMS["initial_relative_weighting"],
#         "procrustes_interval":        OPT_PARAMS["procrustes_interval"],
#         "procrustes_scaling":         OPT_PARAMS["procrustes_scaling"],
#         "save_init_splits":           OPT_PARAMS["save_init_splits"],
#         "verbosity":                  OPT_PARAMS["verbosity"]
#     }

#     if OPT_PARAMS.get("tiny_test", False):
#         valid_params["number_of_particles"] = 32
#         valid_params["optimization_iterations"] = 25

#     if not OPT_PARAMS.get("use_single_scale", 0):
#         valid_params["multiscale"] = 1
#         valid_params["multiscale_particles"] = OPT_PARAMS["multiscale_particles"]

#     for k, v in valid_params.items():
#         parameters.set(k, sw.Variant([v]))

#     project.set_parameters("optimize", parameters)

#     proj_file = os.path.join(output_path, f"{DATASET_NAME}.swproj")
#     project.save(proj_file)

#     color_print("  Lancement de l'optimisation via shapeworks...", Fore.YELLOW)
#     cmd = ['shapeworks', 'optimize', '--progress', '--name', proj_file]
#     subprocess.check_call(cmd, cwd=output_path)

#     args_for_check = type('ArgsForCheck', (object,), {})()
#     args_for_check.tiny_test = OPT_PARAMS.get("tiny_test", False)
#     args_for_check.verify    = False
#     sw.utils.check_results(args_for_check, proj_file)

#     # Créer un dossier pour stocker les meshes alignés
#     aligned_meshes_path = os.path.join(output_path, "aligned_meshes")
#     os.makedirs(aligned_meshes_path, exist_ok=True)

#     for i, subject in enumerate(subjects):
#         mesh = sw.Mesh(subject.get_groomed_filenames()[0])
#         transform = subject.get_groomed_transforms()[0]
#         mesh = mesh.applyRigidTransform(transform)
        
#         basename = os.path.splitext(os.path.basename(shape_filenames[i]))[0]
#         aligned_path = os.path.join(aligned_meshes_path, f"aligned_{basename}.vtk")
#         mesh.write(aligned_path)

#     return os.path.join(output_path, f"{DATASET_NAME}_particles")


# ### 4. PCA et Métriques d'Erreur

# In[20]:


def compute_pca(particles_dir, pca_output_dir):
    color_print("\n--- Step 5. PCA ---", Fore.GREEN, Style.BRIGHT)
    os.makedirs(pca_output_dir, exist_ok=True)

    parts, names = get_particles(particles_dir, "world")
    if parts is None:
        raise ValueError("Aucune particule chargée depuis " + particles_dir)

    n, p, _ = parts.shape
    parts_flat = parts.reshape(n, -1)
    color_print(f"  Forme des particules aplaties : {parts_flat.shape}", Fore.YELLOW)

    pca = PCA(n_components=n - 1)
    pca.fit(parts_flat)
    comps = pca.transform(parts_flat)

    eigvals = pca.explained_variance_
    with open(os.path.join(pca_output_dir, 'eigenvalues.eval'), 'w') as f:
        for ev in eigvals:
            f.write(f"{ev}\n")

    eigenvectors = pca.components_
    eigenvectors_reshaped = eigenvectors.reshape(eigenvectors.shape[0], -1, 3)
    for i, eigvec in enumerate(eigenvectors_reshaped):
        fn = os.path.join(pca_output_dir, f"eigenvector_{i+1}.eig")
        np.savetxt(fn, eigvec, fmt='%f')

    pca_projection = comps[:, :2]
    color_print("  PCA calculée et sauvegardée.", Fore.YELLOW)

    return pca_projection, eigvals, names

def compute_compactness(eigenvalues, threshold=0.95):
    total_var = np.sum(eigenvalues)
    cum_var = np.cumsum(eigenvalues) / total_var
    num_comp = int(np.argmax(cum_var >= threshold) + 1)
    return num_comp, cum_var

def compute_specificity(real_shapes, num_particles, num_samples=1000):
    color_print("  Calcul Specificity...", Fore.YELLOW)
    n, p, dim3 = real_shapes.shape
    d = p * dim3
    real_shapes_2d = real_shapes.reshape(n, d)

    Y = real_shapes_2d.T
    mu = np.mean(Y, axis=1, keepdims=True)
    Yc = Y - mu
    U, S, _ = np.linalg.svd(Yc, full_matrices=False)
    if S[0] < S[-1]:
        S = S[::-1]
        U = np.fliplr(U)

    specifics = np.zeros(n - 1)

    def shape_distance(ptsA, ptsB, pcount):
        A3 = ptsA.reshape(pcount, 3)
        B3 = ptsB.reshape(pcount, 3)
        return np.linalg.norm(A3 - B3, axis=1).sum()

    for m in tqdm(range(1, n), desc="  Specificity modes"):
        epsi = U[:, :m]
        stdevs = np.sqrt(S[:m])
        betas = np.random.randn(m, num_samples)
        for i_mode in range(m):
            betas[i_mode, :] *= stdevs[i_mode]
        synth = epsi @ betas + mu
        min_dists = np.zeros(num_samples)
        for isyn in range(num_samples):
            sy = synth[:, isyn]
            best = 1e15
            for j in range(n):
                dist_j = shape_distance(sy, Y[:, j], num_particles)
                if dist_j < best:
                    best = dist_j
            min_dists[isyn] = best
        specifics[m-1] = np.mean(min_dists) / float(num_particles)

    return specifics

def compute_generalization(real_shapes, num_particles):
    color_print("  Calcul Generalization...", Fore.YELLOW)
    if len(real_shapes.shape) == 3 and real_shapes.shape[2] == 3:
        n, p, dim3 = real_shapes.shape
        d = p * dim3
        real_shapes_2d = real_shapes.reshape(n, d)
    else:
        n, d = real_shapes.shape
        real_shapes_2d = real_shapes

    def shape_distance(ptsA, ptsB, pcount):
        A3 = ptsA.reshape(pcount, 3)
        B3 = ptsB.reshape(pcount, 3)
        return np.linalg.norm(A3 - B3, axis=1).sum()

    P = real_shapes_2d.T
    gens = np.zeros(n - 1)

    for m in range(1, n):
        tot_dist = 0.0
        for leave in range(n):
            Y = np.zeros((P.shape[0], n-1))
            Y[:, :leave] = P[:, :leave]
            Y[:, leave:] = P[:, leave+1:]
            mu = np.mean(Y, axis=1, keepdims=True)
            Yc = Y - mu
            U, S, _ = np.linalg.svd(Yc, full_matrices=False)
            epsi = U[:, :m]

            ytest = P[:, leave:leave+1]
            betas = epsi.T @ (ytest - mu)
            rec = epsi @ betas + mu

            dist = shape_distance(rec, ytest, num_particles) / float(num_particles)
            tot_dist += dist
        gens[m - 1] = tot_dist / float(n)

    return gens

def compute_error_metrics(particles_dir, pca_output_dir, num_particles):
    color_print("\n--- Step 6. Metrics ---", Fore.GREEN, style="")
    real_shapes, real_names = get_particles(particles_dir, "world")
    if real_shapes.size == 0:
        raise ValueError(f"Aucune shape chargée dans {particles_dir}")

    eigenvalues_path = os.path.join(pca_output_dir, 'eigenvalues.eval')
    if not os.path.exists(eigenvalues_path):
        raise FileNotFoundError("Fichier eigenvalues.eval introuvable : " + eigenvalues_path)
    eigenvalues = np.loadtxt(eigenvalues_path)

    c_required, c_variance = compute_compactness(eigenvalues)
    specifics = compute_specificity(real_shapes, num_particles)
    generals = compute_generalization(real_shapes, num_particles)

    metrics = {
        "compactness_required": c_required,
        "cumulative_variance": c_variance.tolist(),
        "specificity": specifics.tolist(),
        "generalization": generals.tolist()
    }
    return metrics


# ### 5. run_optimization_and_analysis

# In[21]:


def run_optimization_and_analysis(run_params, run_index,
                                  shape_seg_list, shape_filenames,
                                  rigid_transforms, groomed_files,
                                  base_output_dir):
    """
    Steps 4,5,6 : Optimization, PCA, Metrics
    """
    color_print(f"\n  >>> RUN {run_index} : Optim + Analyse <<<", Fore.MAGENTA, style="")

    overall_start = time.time()

    run_dir = os.path.join(base_output_dir, f"Run_{run_index}")
    os.makedirs(run_dir, exist_ok=True)
    output_path = os.path.join(run_dir, "OUTPUT")
    os.makedirs(output_path, exist_ok=True)
    pca_out = os.path.join(output_path, "PCA_results")

    # Mise à jour des OPT_PARAMS
    for k,v in run_params.items():
        OPT_PARAMS[k] = v

    step_times = {}

    # Step 4
    t0 = time.time()
    particles_dir = optimize_particles(shape_seg_list, shape_filenames, rigid_transforms, groomed_files, output_path)
    step_times["optimization"] = time.time() - t0

    # Step 5
    t0 = time.time()
    pca_projection, eigvals, shape_names_for_pca = compute_pca(particles_dir, pca_out)
    step_times["pca"] = time.time() - t0

    # Step 6
    t0 = time.time()
    n_parts = OPT_PARAMS.get("number_of_particles", 128)
    metrics = compute_error_metrics(particles_dir, pca_out, n_parts)
    step_times["error_metrics"] = time.time() - t0

    overall_time = time.time() - overall_start
    color_print(f"  RUN {run_index} terminé en {overall_time:.2f}s", Fore.MAGENTA)

    # On remet les OPT_PARAMS comme avant si besoin (pas forcément)
    return {
        "pca_projection": pca_projection,
        "pca_shape_names": shape_names_for_pca,
        "metrics": metrics,
        "params": run_params,
        "step_times": step_times,
        "total_execution_time": overall_time
    }


# ### 6. Création du Fichier Excel

# In[22]:


def mm_ss_format(seconds):
    mm = int(seconds // 60)
    ss = int(seconds % 60)
    return f"{mm}:{ss:02d}"

def _plot_metric_curve(data, title, ylabel, run_idx, outdir, figsize=(5, 3)):
    os.makedirs(outdir, exist_ok=True)
    fn = f"{title.replace(' ', '_')}_Run_{run_idx}.png"
    image_path = os.path.join(outdir, fn)

    fig, ax = plt.subplots(figsize=figsize)
    ax.plot(range(1, len(data) + 1), data, marker='o')
    ax.set_title(title)
    ax.set_xlabel("Number of Modes")
    ax.set_ylabel(ylabel)
    ax.grid(True)

    plt.savefig(image_path, dpi=130, bbox_inches="tight")
    plt.close(fig)

    return image_path

def _plot_pca_scatter(pc, shape_names, run_idx, outdir):
    """
    Scatter plot (PC1 vs PC2), indexés 1..N + tableau associant index=>shape_name
    """
    os.makedirs(outdir, exist_ok=True)
    fn_img = f"PCA_Scatter_Run_{run_idx}.png"
    image_path = os.path.join(outdir, fn_img)

    fig, ax = plt.subplots(figsize=(4,4))
    x = pc[:,0]
    y = pc[:,1]
    ax.scatter(x, y, s=30, c='blue')

    for i, (xx, yy) in enumerate(zip(x,y)):
        ax.text(xx, yy, str(i+1), fontsize=8, color='red')

    ax.set_title("PC1 vs PC2 Scatter")
    ax.set_xlabel("PC1")
    ax.set_ylabel("PC2")
    ax.grid(True)

    plt.tight_layout()
    plt.savefig(image_path, dpi=130)
    plt.close(fig)

    df_map = pd.DataFrame({
        "Index": list(range(1, len(shape_names)+1)),
        "ShapeName": shape_names
    })

    return image_path, df_map


def save_results_to_excel(all_results, excel_filename, grooming_keys, optimization_keys):

    color_print("\nSauvegarde des résultats dans le fichier Excel...", Fore.CYAN)

    wide_rows = []
    columns_order = (list(grooming_keys) + list(optimization_keys) +
                     ["time_grooming", "time_rigid", "time_optimization", "time_total",
                      "compactness_95", "final_specificity_error", "final_generalization_error"]
                    )

    for i, res in enumerate(all_results, start=1):
        # on merge grooming_params + res["params"]
        row_dict = {}

        # grooming
        groom_p = res.get("grooming_params", {})
        for gk in grooming_keys:
            row_dict[gk] = groom_p.get(gk, None)

        # optimization
        optim_p = res["params"]
        for ok in optimization_keys:
            row_dict[ok] = optim_p.get(ok, None)

        # times
        t_groom = res.get("time_grooming", 0)
        t_rigid = res.get("time_rigid", 0)

        row_dict["time_grooming"] = mm_ss_format(t_groom)
        row_dict["time_rigid"]    = mm_ss_format(t_rigid)

        st = res.get("step_times", {})
        t_opt = st.get("optimization", 0)
        t_tot = res.get("total_execution_time", 0)
        row_dict["time_optimization"] = mm_ss_format(t_opt)
        row_dict["time_total"]        = mm_ss_format(t_tot)

        # metrics
        mets = res["metrics"]
        row_dict["compactness_95"] = mets["compactness_required"]
        spec = mets["specificity"]
        row_dict["final_specificity_error"] = ( spec[-1] if len(spec)>0 else None )
        gen = mets["generalization"]
        row_dict["final_generalization_error"] = ( gen[-1] if len(gen)>0 else None )

        # on l'ajoute
        wide_rows.append(row_dict)

    df_wide = pd.DataFrame(wide_rows, columns=columns_order)
    # on va rename l'index => "Run_1", "Run_2", ...
    run_names = [f"Run_{i}" for i in range(1, len(all_results)+1)]
    df_wide.index = run_names

    # on transpose
    df_tall = df_wide.transpose()

    # on fait un reset => la 1re col "Field"
    df_tall.insert(0, "Field", df_tall.index)
    df_tall.reset_index(drop=True, inplace=True)

    # 3) On crée le workbook
    writer = pd.ExcelWriter(excel_filename, engine="openpyxl")

    # On écrit df_tall dans "RESULTS"
    df_tall.to_excel(writer, sheet_name="RESULTS", index=False)

    ws_results = writer.book["RESULTS"]

    # -- Mieux présenter la page : 
    ws_results.column_dimensions["A"].width = 25  # la col "Field"
    # on met un width un peu plus large pour chaque col "Run_i"
    for col_i in range(2, len(run_names)+2):
        colL = ws_results.cell(row=1, column=col_i).column_letter
        ws_results.column_dimensions[colL].width = 14

    # 4) Placer TOUS LES PARAMS en colonnes D/E, en dessous ?
    #    On va récolter tout l'univers (Grooming + Optim) => faire un listing trié
    all_groom_keys = ["ANTIALIAS_ITERATIONS", "ISO_SPACING", "PAD_SIZE",
                      "PAD_VALUE", "ISO_VALUE", "ICP_ITERATIONS"]
    all_optim_keys = sorted(OPT_PARAMS.keys())
    # On va lister grooming + optim keys, 
    row_start_params = df_tall.shape[0] + 3
    row_cur = row_start_params
    ws_results.cell(row=row_cur, column=4, value="---- GROOMING PARAMS ----")
    row_cur += 1
    for gk in all_groom_keys:
        val = None
        if gk == "ANTIALIAS_ITERATIONS":
            val = ANTIALIAS_ITERATIONS
        elif gk == "ISO_SPACING":
            val = str(ISO_SPACING)
        elif gk == "PAD_SIZE":
            val = PAD_SIZE
        elif gk == "PAD_VALUE":
            val = PAD_VALUE
        elif gk == "ISO_VALUE":
            val = ISO_VALUE
        elif gk == "ICP_ITERATIONS":
            val = ICP_ITERATIONS
        else:
            val = None

        ws_results.cell(row=row_cur, column=4, value=gk)
        ws_results.cell(row=row_cur, column=5, value=str(val))
        row_cur += 1

    row_cur += 1
    ws_results.cell(row=row_cur, column=4, value="---- OPTIM PARAMS ----")
    row_cur += 1
    for ok in all_optim_keys:
        val = OPT_PARAMS.get(ok, None)
        ws_results.cell(row=row_cur, column=4, value=ok)
        ws_results.cell(row=row_cur, column=5, value=str(val))
        row_cur += 1

    # 5) Feuilles "Run_i"
    for i, res in enumerate(all_results, start=1):
        sheet_name = f"Run_{i}"
        dummy_df = pd.DataFrame()
        dummy_df.to_excel(writer, sheet_name=sheet_name, index=False)

        ws_run = writer.book[sheet_name]

        # Graphs
        run_dir = os.path.join(BASE_OUTPUT_DIR, f"Run_{i}")
        plot_dir = os.path.join(run_dir, "plots")
        os.makedirs(plot_dir, exist_ok=True)

        mets = res["metrics"]
        cvar = mets["cumulative_variance"]
        specificity_data = mets["specificity"]
        general_data = mets["generalization"]

        compactness_img = _plot_metric_curve(cvar, "Compactness", "Variance", i, plot_dir)
        specificity_img = _plot_metric_curve(specificity_data, "Specificity Error", "Error", i, plot_dir)
        general_img     = _plot_metric_curve(general_data, "Generalization Error", "Error", i, plot_dir)

        row_img1 = 1
        row_img2 = 12
        row_img3 = 23

        imgA = ExcelImage(compactness_img); imgA.width, imgA.height = 310, 180
        imgB = ExcelImage(specificity_img); imgB.width, imgB.height = 310, 180
        imgC = ExcelImage(general_img);     imgC.width, imgC.height = 310, 180

        ws_run.add_image(imgA, f"A{row_img1}")
        ws_run.add_image(imgB, f"A{row_img2}")
        ws_run.add_image(imgC, f"A{row_img3}")

        ws_run.cell(row=row_img1+9, column=1, value=f"Composantes pour 95%: {mets['compactness_required']}")
        if len(specificity_data) > 0:
            ws_run.cell(row=row_img2+9, column=1, 
                        value=f"Specificity final error: {specificity_data[-1]:.4f}")
        if len(general_data) > 0:
            ws_run.cell(row=row_img3+9, column=1, 
                        value=f"Generalization final error: {general_data[-1]:.4f}")

        pc = res["pca_projection"]
        shape_names = res.get("pca_shape_names", [])
        # On met PC1/PC2
        start_row_pc = 34
        pc1_pc2_df = pd.DataFrame(pc, columns=["PC1","PC2"])
        pc1_pc2_df.to_excel(writer, sheet_name=sheet_name, startrow=start_row_pc, index=False)

        scatter_img, df_map_idx = _plot_pca_scatter(pc, shape_names, i, plot_dir)
        ws_run.add_image(ExcelImage(scatter_img), "L1")

        df_map_idx.to_excel(writer, sheet_name=sheet_name, startrow=start_row_pc+5, startcol=10, index=False)

    writer.close()
    color_print(f"Fichier Excel sauvegardé : {excel_filename}", Fore.CYAN)


#  ## 6. Script Principal (Double Boucle : Grooming & Optim)

# In[23]:


def main():
    global_start = time.time()
    all_results = []
    run_counter = 0

    grooming_dir = os.path.join(BASE_OUTPUT_DIR, "GROOMING")
    os.makedirs(grooming_dir, exist_ok=True)

    # Constructions de combinaisons de grooming
    grooming_keys = list(GRID_GROOMING.keys())
    grooming_values = [GRID_GROOMING[k] for k in grooming_keys]
    if not grooming_keys:
        grooming_combos = [{}]
    else:
        grooming_combos = []
        for combo_g in itertools.product(*grooming_values):
            d = {}
            for k, val in zip(grooming_keys, combo_g):
                d[k] = val
            grooming_combos.append(d)

    # Constructions de combinaisons d'optim
    optim_keys = list(GRID_OPTIMIZATION.keys())
    optim_values = [GRID_OPTIMIZATION[k] for k in optim_keys]
    optim_combos = []
    for combo_o in itertools.product(*optim_values):
        dd = {}
        for k, val in zip(optim_keys, combo_o):
            dd[k] = val
        optim_combos.append(dd)

    # On boucle
    for ig, groom_params in enumerate(grooming_combos, start=1):
        if len(grooming_combos) > 1:
            color_print(f"\n================= GROOMING Variation {ig}/{len(grooming_combos)} =================", Fore.BLUE, Style.BRIGHT)
            color_print(str(groom_params), Fore.BLUE)

        # Rerun grooming
        shape_seg_list, shape_filenames, dataset_ids, shape_names, \
        rigid_transforms, groomed_files, t_groom, t_rigid = run_preprocessing(
            DATASET_PATHS, SHAPE_EXT, groom_params,
            base_dir=os.path.join(grooming_dir, f"Groom_{ig}")
        )

        # On va stocker t_groom, t_rigid => on le recopie ensuite sur chaque run
        for io, optim_params in enumerate(optim_combos, start=1):
            run_counter += 1
            print("\n-----------------------------------------------------")
            color_print(f"\n  >>> RUN {run_counter} / G={ig}, O={io} <<<", Fore.MAGENTA, Style.BRIGHT)
            color_print("   Optim params:" + str(optim_params), Fore.MAGENTA)

            # Exécuter
            out = run_optimization_and_analysis(
                run_params=optim_params,
                run_index=run_counter,
                shape_seg_list=shape_seg_list,
                shape_filenames=shape_filenames,
                rigid_transforms=rigid_transforms,
                groomed_files=groomed_files,
                base_output_dir=BASE_OUTPUT_DIR
            )
            # On ajoute l'info grooming
            out["grooming_params"] = groom_params
            out["time_grooming"]   = t_groom
            out["time_rigid"]      = t_rigid
            all_results.append(out)

    # Export excel
    excel_filename = os.path.join(BASE_OUTPUT_DIR, DATASET_NAME + ".xlsx")
    save_results_to_excel(all_results, excel_filename, grooming_keys, optim_keys)

    total_time = time.time() - global_start
    color_print(f"\nPipeline terminée en {total_time:.2f}s (global).", Fore.GREEN, Style.BRIGHT)

    # Compress Output Pipeline
    import shutil
    
    # Dossier à zipper
    dossier_source = '/home/jupyter-gossard/Code_PFE/OUTPUT_PIPELINE'
    # Nom de sortie sans extension
    nom_zip = '/home/jupyter-gossard/Code_PFE/'+ DATASET_NAME  # ne pas ajouter '.zip'
    color_print(f"\n Compression Output Pipeline {DATASET_NAME}", Fore.GREEN, Style.BRIGHT)
    # Création du zip
    shutil.make_archive(nom_zip, 'zip', dossier_source)


if __name__ == "__main__":
    main()

